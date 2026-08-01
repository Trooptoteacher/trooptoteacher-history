# -*- coding: utf-8 -*-
"""Build the course-wide primary-source SOURCING CHECKLIST workbook (.xlsx) from
the compiled, verified link data (primary_source_sourcing.json).

Each row = one primary source to download for the Government & Civics course:
unit, standard, the work, type, creator/year, repository, rights status,
a clickable catalog/page link, a clickable direct-download link, the target
filename, and a Status column the sourcing team fills in.
"""
import os, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "primary_source_sourcing.json")
OUT = os.path.join(HERE, "Government_Hack_Primary_Source_Sourcing_List.xlsx")

rows = json.load(open(DATA))["sources"]

NAVY = "1B2A4A"; RED = "B22234"; GOLD = "C89B3C"; CREAM = "F7F5EF"
UNIT_TINT = {1:"EEF1F7",2:"FBEEEF",3:"FAF3E2",4:"EAF2EC",5:"EEF1F7",6:"FBEEEF",7:"FAF3E2"}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Primary Sources"

cols = ["Unit","Standard","Standard Title","Primary Source (work)","Type",
        "Creator","Year","Repository","Rights / Public-use","Verified?",
        "Catalog / Page link","Direct download","Save-as filename","Status"]
ws.append(cols)

hdr_fill = PatternFill("solid", fgColor=NAVY)
hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="CBD2DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c in range(1, len(cols)+1):
    cell = ws.cell(row=1, column=c)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 30

link_font = Font(name="Calibri", color="0563C1", underline="single", size=10)
base_font = Font(name="Calibri", size=10)
r = 2
for s in rows:
    unit = s["unit"]
    page = s.get("page_url") or ""
    direct = s.get("direct_file_url") or ""
    verified = "YES" if s.get("verified") else "check"
    vals = [unit, s["standard"], s.get("standard_title",""), s.get("work_title",""),
            s.get("type",""), s.get("creator",""), s.get("year",""), s.get("repository",""),
            s.get("rights",""), verified,
            "Open catalog" if page else (s.get("search_hint","") or "—"),
            "Download" if direct else "—",
            s.get("filename",""), "To source"]
    ws.append(vals)
    tint = PatternFill("solid", fgColor=UNIT_TINT.get(unit, "FFFFFF"))
    for c in range(1, len(cols)+1):
        cell = ws.cell(row=r, column=c)
        cell.border = border; cell.font = base_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if c in (1,2,5,7,10,14):
            cell.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
        if c == 1:
            cell.fill = PatternFill("solid", fgColor=GOLD); cell.font = Font(bold=True, size=10)
        elif c in (3,):
            cell.fill = tint
    # hyperlinks
    if page:
        lc = ws.cell(row=r, column=11); lc.hyperlink = page; lc.font = link_font
    if direct:
        dc = ws.cell(row=r, column=12); dc.hyperlink = direct; dc.font = link_font
    # verified coloring
    vc = ws.cell(row=r, column=10)
    vc.fill = PatternFill("solid", fgColor=("EAF2EC" if s.get("verified") else "FBEEEF"))
    r += 1

widths = [6,10,26,34,11,20,7,26,30,10,16,13,26,12]
for i,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{r-1}"

# summary sheet
s2 = wb.create_sheet("How to use")
notes = [
 ["Government Hack — Primary Source Sourcing List"],
 [""],
 ["Purpose","One row per primary source to download for the U.S. Government & Civics course (GC.01–GC.35)."],
 ["Public use","All listed works are intended to be public domain: U.S./TN government works, U.S. Supreme Court opinions, works published before 1929, or works whose creator died over 95 years ago. Confirm each digital reproduction's terms at the repository before publishing."],
 ["Verified?","YES = a research pass found and opened a live catalog/Commons page for this exact item. 'check' = not yet confirmed; use the search hint in the link column."],
 ["Catalog link","Opens the item's page at Wikimedia Commons / Library of Congress / National Archives — use its Download button."],
 ["Direct download","When present, links straight to the full-resolution image file."],
 ["Save-as filename","Save each file with EXACTLY this name and drop it in the course image bank; the deck/workbook builds then embed it automatically."],
 ["Rights caution","Modern (post-1928) news photos and cartoons are usually copyrighted — for 20th/21st-century landmark cases we point to public-domain government documents (amendment texts, court filings) or federal-government photographs instead."],
]
for row in notes: s2.append(row)
s2.column_dimensions["A"].width = 16; s2.column_dimensions["B"].width = 100
s2["A1"].font = Font(bold=True, size=14, color=NAVY)
for rr in range(3, len(notes)+1):
    s2.cell(row=rr, column=1).font = Font(bold=True)
    s2.cell(row=rr, column=2).alignment = Alignment(wrap_text=True, vertical="top")

wb.save(OUT)
print("wrote", OUT, "|", len(rows), "sources")
