import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT="/tmp/claude-0/-home-user-trooptoteacher-history/c27aff16-3c79-5004-bf4a-470896c27424/scratchpad/crosswalks"
COURSES=[
 ("World History (High School)","courses/world-history/standards/world-history-standards.json","repo-verified"),
 ("Tennessee History","courses/tennessee-history/standards/tennessee-history-standards.json","repo-verified"),
 ("US History — Grade 8","courses/us-history-grade8/standards/us-history-grade8-standards.json","GitHub-sourced — VERIFY vs official TDOE PDF"),
 ("World History & Geography — Grade 7","courses/world-history-grade7/standards/world-history-grade7-standards.json","GitHub-sourced+cross-checked — exemplar lists omitted; VERIFY vs official PDF"),
]
COLS=["Unit","Standard","Standard Text (verbatim)","Geographic Feature?","TN Connection?",
 "Primary Source (work)","Type","Creator","Year","Repository","Rights / License","Commercial Use OK?",
 "Verified?","Catalog URL","Direct download URL","Save-as filename","Citation","MIME type","Width","Height","SHA-256","Status"]

NAVY="0A1F3C"; GOLD="C89B3C"; CREAM="F7F5EF"; GEO="DCE6F1"; TNC="EAF3E4"
hdr_fill=PatternFill("solid",fgColor=NAVY); hdr_font=Font(color="FFFFFF",bold=True,size=10)
thin=Side(style="thin",color="D9D9D9"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
wrap=Alignment(wrap_text=True,vertical="top")

RULES=[
 "PRIMARY-SOURCE SOURCING RULES (from the existing History Hack Sourcing List schema):",
 "",
 "APPROVED REPOSITORIES ONLY: Library of Congress, National Archives (NARA), Office of the Historian,",
 "National Park Service (NPS), Smithsonian, Tennessee Encyclopedia / TeVA, other *.gov / *.edu.",
 "Wikimedia Commons / DPLA allowed ONLY for genuinely public-domain works, and PREFER the original",
 "repository (swap to LoC/NARA/TN State when the same item exists there). NEVER Britannica or Wikipedia.",
 "",
 "COMMERCIAL USE: every source must be cleared for commercial use — Public Domain, CC0, CC-BY, or US-gov work.",
 "Record the exact rights statement in 'Rights / License' and set 'Commercial Use OK?' = YES only when confirmed.",
 "",
 "2-4 primary sources per standard (one per row; repeat the Standard code across its rows).",
 "Save each image as {STANDARD}_short-name.ext so the build embeds it automatically.",
 "'Verified?' = YES only after the Catalog URL AND Direct download URL have been opened and confirmed to resolve.",
 "",
 "=========================================================================================",
 "STATUS OF THIS FILE (2026-08-02): SCAFFOLD ONLY — verified sources NOT yet populated.",
 "",
 "WHY: This build environment's egress policy BLOCKS the archive hosts (loc.gov, catalog.archives.gov,",
 "si.edu, tn.gov, archive.org all return 403 policy denial; only GitHub is reachable). Verified catalog/",
 "download links and license checks CANNOT be produced here without routing around org policy (which is",
 "prohibited). No links were guessed or fabricated.",
 "",
 "TO FINISH: allowlist the archive hosts in the environment's network policy (or run the sourcing pass in an",
 "environment that can reach them). Then the source columns fill per the rules above, 2-4 verified rows/standard.",
 "",
 "STANDARDS CAVEAT: see each course's source provenance. World History & Tennessee History standards are",
 "repo-verified. Grade 7 & Grade 8 standards were reconstructed from GitHub mirrors + WebSearch cross-check",
 "(official tn.gov PDF was unreachable) — VERIFY verbatim text against the official TDOE PDF before publishing.",
]

def tn_flag(course, text):
    if course=="Tennessee History": return "Yes"
    return "Yes" if ("tennessee" in text.lower() or "tenn." in text.lower()) else ""

built=[]
for title, path, prov in COURSES:
    if not os.path.exists(path): 
        print("skip (missing):",path); continue
    d=json.load(open(path)); st=d.get("standards",[])
    if not st: 
        print("skip (0 standards):",path); continue
    wb=Workbook()
    ws=wb.active; ws.title="Crosswalk"
    # header
    for c,name in enumerate(COLS,1):
        cell=ws.cell(1,c,name); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=Alignment(wrap_text=True,vertical="center"); cell.border=border
    ws.freeze_panes="A2"
    r=2
    for s in st:
        geo = "Yes" if s.get("geo") else ""
        vals=[s.get("cluster",""), s.get("code",""), s.get("text",""), geo, tn_flag(title,s.get("text","")),
              "","","","","","","","","","","","","","","","",
              "PENDING — verified source blocked by egress policy (see README tab)"]
        for c,v in enumerate(vals,1):
            cell=ws.cell(r,c,v); cell.alignment=wrap; cell.border=border
            if c==4 and geo: cell.fill=PatternFill("solid",fgColor=GEO)
            if c==5 and vals[4]=="Yes": cell.fill=PatternFill("solid",fgColor=TNC)
        r+=1
    widths=[26,10,60,14,13,26,12,16,8,20,22,15,10,26,26,22,34,12,8,8,20,44]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for row in range(2,r): ws.row_dimensions[row].height=42
    # README tab
    rd=wb.create_sheet("README & Rules")
    rd.column_dimensions["A"].width=100
    rd.cell(1,1,f"{title} — Primary Source Crosswalk").font=Font(bold=True,size=14,color=NAVY)
    rd.cell(2,1,f"Standards: {len(st)}  |  Geo-flagged: {sum(1 for s in st if s.get('geo'))}  |  Provenance: {prov}").font=Font(italic=True,size=10)
    for i,line in enumerate(RULES,4):
        c=rd.cell(i,1,line); c.alignment=Alignment(wrap_text=True,vertical="top")
        if line.endswith(":") or line.startswith("STATUS") or line.startswith("WHY") or line.startswith("TO FINISH") or line.startswith("STANDARDS CAVEAT"): c.font=Font(bold=True)
    fn=f"{OUT}/{title.replace(' ','_').replace('&','and').replace('(','').replace(')','').replace('—','-')}_Primary_Source_Crosswalk.xlsx"
    wb.save(fn); built.append((title,len(st),fn)); print(f"built: {len(st):3d} rows -> {os.path.basename(fn)}")

print("\n== summary ==")
for t,n,f in built: print(f"  {n:3d} standards  {t}")
