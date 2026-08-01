#!/usr/bin/env python3
"""
Ingest a course Primary-Source Sourcing List (.xlsx from Drive) into the course
image asset catalog (`<course>_images.json`) — so cited/licensed assets flow into
the build instead of being re-sourced.

The sourcing spreadsheets (e.g. Government_Hack_Primary_Source_Sourcing_List_
Completed.xlsx) carry a "Download Manifest" sheet with verified, adoption-grade
columns:
  Standard | Work | Save-as filename | Status | Catalog URL | Direct download URL |
  Verified title | Verified creator | Verified date | Repository | Rights / License |
  Citation | MIME type | Width | Height | File size (bytes) | SHA-256

This tool maps each row to the `anchor` asset record used by the workbook/deck
builders, keyed by standard code:
  { file, medium, title, creator, year, citation, rights, rightsUrl(catalog),
    mime, w, h, sha256, source_repository }

Output merges into (does not clobber) an existing images JSON.

Usage:
  python3 ingest_sourcing_list.py <sourcing.xlsx> <out_images.json> [--sheet "Download Manifest"] [--prefix GC]
"""
import sys, os, json
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

argv = sys.argv[1:]
flags = {a.split("=")[0].lstrip("-"): (a.split("=", 1)[1] if "=" in a else True)
         for a in argv if a.startswith("--")}
pos = [a for a in argv if not a.startswith("--")]
if len(pos) < 2:
    sys.exit("Usage: ingest_sourcing_list.py <sourcing.xlsx> <out_images.json> [--sheet NAME] [--prefix GC]")
XLSX, OUT = pos[0], pos[1]
SHEET = flags.get("sheet", "Download Manifest")
PREFIX = flags.get("prefix")  # optional: only ingest rows whose Standard starts with this

# medium inference from filename/work
def infer_medium(fname, work):
    f = (fname or "").lower(); w = (work or "").lower()
    if f.endswith(".svg") or "seal" in w or "flag" in w or "emblem" in w: return "emblem"
    for k in ("map",):
        if k in w or k in f: return "map"
    for k in ("cartoon", "gerry-mander", "pillars", "looking glass", "panic", "press"):
        if k in w: return "cartoon"
    for k in ("portrait",):
        if k in w: return "portrait"
    for k in ("photo", "photograph", "parade", "capitol", "building", "immigrants", "incarceration"):
        if k in w: return "photo"
    if any(k in w for k in ("declaration", "constitution", "amendment", "act", "federalist",
                            "bill of rights", "petition", "resolution", "seal")): return "document"
    return "image"

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
if SHEET not in wb.sheetnames:
    # fall back to the sheet that has a "SHA-256" or "Citation" header
    SHEET = next((s for s in wb.sheetnames
                  if any("sha-256" in str(c.value).lower() or "citation" == str(c.value).lower()
                         for c in next(wb[s].iter_rows(min_row=1, max_row=1)))), wb.sheetnames[-1])
ws = wb[SHEET]

rows = list(ws.iter_rows(values_only=True))
# locate the header row (contains "Standard" and "Citation")
hdr_idx = next((i for i, r in enumerate(rows)
                if r and any(str(c).strip().lower() == "standard" for c in r if c)
                and any("citation" in str(c).strip().lower() for c in r if c)), 0)
header = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
def col(name):
    for i, h in enumerate(header):
        if h.lower() == name.lower(): return i
    return None
ci = {k: col(k) for k in ["Standard", "Work", "Save-as filename", "Catalog URL",
                          "Verified title", "Verified creator", "Verified date", "Repository",
                          "Rights / License", "Citation", "MIME type", "Width", "Height", "SHA-256"]}

catalog = {}
skipped = 0
for r in rows[hdr_idx + 1:]:
    if not r or ci["Standard"] is None: continue
    std = r[ci["Standard"]]
    if not std or not str(std).strip(): continue
    std = str(std).strip()
    if PREFIX and not std.startswith(PREFIX): continue
    g = lambda k: (r[ci[k]] if ci[k] is not None and ci[k] < len(r) else None)
    fname = g("Save-as filename")
    rec = {
        "file": str(fname).strip() if fname else "",
        "medium": infer_medium(str(fname or ""), str(g("Work") or g("Verified title") or "")),
        "title": str(g("Verified title") or g("Work") or "").strip(),
        "creator": str(g("Verified creator") or "").strip(),
        "year": str(g("Verified date") or "").strip(),
        "citation": str(g("Citation") or "").strip(),
        "rights": str(g("Rights / License") or "").strip(),
        "rightsUrl": str(g("Catalog URL") or "").strip(),
        "source_repository": str(g("Repository") or "").strip(),
        "mime": str(g("MIME type") or "").strip(),
        "w": g("Width"), "h": g("Height"),
        "sha256": str(g("SHA-256") or "").strip(),
    }
    if not rec["citation"] or not rec["rights"]:
        skipped += 1
    catalog[std] = {"anchor": rec}

# merge into existing images JSON (do not clobber other keys)
existing = {}
if os.path.exists(OUT):
    existing = json.load(open(OUT, encoding="utf8"))
for std, rec in catalog.items():
    existing.setdefault(std, {}).update(rec)
json.dump(existing, open(OUT, "w", encoding="utf8"), ensure_ascii=False, indent=2)

print(f"Ingested {len(catalog)} assets from '{SHEET}' -> {OUT}")
if skipped:
    print(f"  WARNING: {skipped} rows missing citation or rights/license")
# quick repository breakdown (for the TDOE approved-source check)
from collections import Counter
repos = Counter(v["anchor"]["source_repository"] for v in catalog.values())
print("  repositories: " + ", ".join(f"{k or '(blank)'}={n}" for k, n in repos.most_common()))
