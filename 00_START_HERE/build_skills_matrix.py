# -*- coding: utf-8 -*-
"""Master Skills Matrix — every History Hack / TroopToTeacher skill, where it lives,
and the exact prompt to trigger it. America 250 styling; frozen header + autofilter."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F3A5F"; RED="B22234"; GOLD="C9A227"; CREAM="F8F5EF"; LIGHT="EEF1F7"; WHITE="FFFFFF"

# (Category, Skill, One-job description, Path, "Say this to trigger it")
ROWS = [
 ("★ START HERE","history-hack-platinum-standard","Mission, Gold-vs-Platinum tier, future-ready principle, and the decision rule (100% alignment → Schedule F → adoption). Auto-loaded at the start of any build.",".claude/skills/history-hack-platinum-standard/","\"Build this to the Platinum standard\"  (or /history-hack-platinum-standard)"),

 ("Build — units & courses","history-hack-new-course-builder","Cradle-to-grave NEW-COURSE orchestrator (Government, Economics, World History, Geography…). Stands up a whole course edition feature-for-feature with US History.",".claude/skills/history-hack-new-course-builder/","\"Build a new [Economics] course edition from these standards\"  ← THE NEW-COURSE PROMPT"),
 ("Build — units & courses","history-hack-platinum-unit-builder","Orchestrator for a full Course Standard unit set (workbook + teacher guide + student & teacher decks). Invokes the engine + gates.",".claude/skills/history-hack-platinum-unit-builder/","\"Build the Unit 3 Course Standard set\"  /  \"Make Unit 3 match the Unit 6 platinum format\""),
 ("Build — units & courses","history-hack-unit-content-build","The 7-activity content ENGINE — student workbook: guided Cornell, back-page UDL/MTSS supports, Future Ready, Lenses, deck-keying, five editions.",".claude/skills/history-hack-unit-content-build/","\"Build/finish/QA the Unit N student workbook (7-activity cycle, guided Cornell, supports)\""),
 ("Build — units & courses","history-hack-dbq-workbook","Standalone DBQ / primary-source SKU (one investigation question, HIPPO/OPTIC, essay + rubric). A different product from the unit workbook.",".claude/skills/history-hack-dbq-workbook/","\"Build a standalone DBQ workbook / primary-source packet for US.NN\""),
 ("Build — units & courses","history-hack-graphic-organizer-workbook","Reproducible graphic-organizer toolkit (Venn, T-chart, matrix, timeline, Frayer, CER, HIPPO, TN Connection) + per-standard best-fits.",".claude/skills/history-hack-graphic-organizer-workbook/","\"Build the Unit N graphic organizer toolkit\"  /  \"Make Unit N organizers match Unit 1\""),

 ("Decks","history-hack-tcap-deck-builder","Teacher (LECTURE) deck .pptx + printable answer-key PDF + usage guide. Per-standard lesson arc, PD imagery, EN/ES word wall.",".claude/skills/history-hack-tcap-deck-builder/","\"Build the Unit N teacher (lecture) TCAP deck + answer key\""),
 ("Decks","history-hack-lean-deck-builder","Student (REVIEW/LEAN) deck .pptx — assertion-evidence, Source It First, Three Perspectives + answer key + web viewer manifest.",".claude/skills/history-hack-lean-deck-builder/","\"Build the Unit N student (lean/review) deck\""),

 ("Packaging & gap-fill","history-hack-poster-packet-builder","24×36 wall posters + Letter station packets + teacher guides + assembled bundle PDFs.",".claude/skills/history-hack-poster-packet-builder/","\"Build the Unit N poster packet / wall set\""),
 ("Packaging & gap-fill","us-history-hack-packet-builder","TpT / for-sale packaging — front/back matter, brand strip, copyright, encryption.",".claude/skills/us-history-hack-packet-builder/","\"Package Unit N for sale (TpT front/back matter, brand, copyright)\""),
 ("Packaging & gap-fill","magic-school-created-worksheets","Copy-ready MagicSchool prompt packets for genuine content gaps.",".claude/skills/magic-school-created-worksheets/","\"Create MagicSchool worksheet prompts for [gap]\""),
 ("Packaging & gap-fill","history-hack-district-edition-restructure","App-side District Edition / product-edition isolation and route protection.",".claude/skills/history-hack-district-edition-restructure/","\"Work on the District Edition / product-edition isolation in the app\""),

 ("Content & design","instructional-design-specialist","Print-first pedagogical architecture of lessons/units/sequences + assessment design (UDL 3.0/CER/C3 → 7-activity flow).",".claude/skills/instructional-design-specialist/","\"Design the lesson/unit structure + assessments for US.NN\""),
 ("Content & design","learning-experience-designer","Interactive digital/app student UX — interaction patterns, journeys, gamification, microlearning.",".claude/skills/learning-experience-designer/","\"Design the app/web student UX / gamification for [feature]\""),
 ("Content & design","udl-cast-expert","UDL 3.0 / CAST design authority — the single owner of UDL principle depth.",".claude/skills/udl-cast-expert/","\"Apply UDL 3.0 / CAST to this activity\""),
 ("Content & design","tt-education-research-team","ESSA evidence tiering / research foundations for a strategy.",".claude/skills/tt-education-research-team/","\"Give me the ESSA evidence tier / research base for [strategy]\""),
 ("Content & design","tn-content-specialist","TN-standards content AUTHORING — narrative prose, lesson narratives (Reconstruction → Modern).",".claude/skills/tn-content-specialist/","\"Write the TN-standards narrative/content for US.NN\""),
 ("Content & design","spaced-repetition-engine","Spiral/retrieval SCHEDULING algorithm (SM-2 / Leitner / forgetting curve). The book seeds; this owns the schedule.",".claude/skills/spaced-repetition-engine/","\"Schedule the spaced retrieval / spiral review for Unit N\""),

 ("Assessment","tn-assessment-specialist","ALL TCAP assessment items + tests — every type, full psychometrics (DOK/Hess/IRT), TCAP JSON, QC + quick-quiz modes.",".claude/skills/tn-assessment-specialist/","\"Write TCAP items / a unit test for US.NN (DOK, IRT, TCAP JSON)\""),

 ("Gates & QC","history-hack-lesson-flow-qc","Release gate: workbook↔deck EXACT-slide alignment, DI parity, vocab-first, 'can a student follow along?'",".claude/skills/history-hack-lesson-flow-qc/","\"QC the Unit N workbook↔deck alignment / student journey\""),
 ("Gates & QC","history-hack-text-integrity-qc","Release gate: no truncated / clipped / placeholder text in decks + workbooks.",".claude/skills/history-hack-text-integrity-qc/","\"Run the text-integrity QC on Unit N\""),
 ("Gates & QC","accessibility-qc-agent","Final WCAG 2.2 AA / 508 / ADA Title II gate (code + content) before human release.",".claude/skills/accessibility-qc-agent/","\"Run the final accessibility gate on [artifact]\""),
 ("Gates & QC","ell-bilingual-review-specialist","EN/ES parity + ELL scaffolding QC (WIDA / ELPA21).",".claude/skills/ell-bilingual-review-specialist/","\"Review the EN/ES + ELL scaffolding on Unit N\""),
 ("Gates & QC","history-hack-unit-qc","End-to-end web-app unit QC workflow (18-item checklist); orchestrates the gates.",".claude/skills/history-hack-unit-qc/","\"Run the unit QC workflow on Unit N / Schedule F prep\""),
 ("Gates & QC","wcs-app-approval-qc","District-submission readiness — Williamson County (WCS) app-approval packet.",".claude/skills/wcs-app-approval-qc/","\"Prep the WCS app-approval packet\""),
 ("Gates & QC","historian-factcheck-agent","Primary-source fact-check (Policy 2.600) — every date/number/name/statute verified.",".claude/skills/historian-factcheck-agent/","\"Fact-check Unit N content against primary sources\""),
 ("Gates & QC","tn-textbook-adoption-agent","TDOE Schedule F / adoption panel review; submission-package prep.",".claude/skills/tn-textbook-adoption-agent/","\"Run the TDOE Schedule F / adoption review on Unit N\""),
 ("Gates & QC","copyright-integrity-accreditation","IP / licensing / FERPA-COPPA / ToS / OSS-license review.",".claude/skills/copyright-integrity-accreditation/","\"Review IP / licensing / FERPA-COPPA for [asset]\""),
 ("Gates & QC","edtech-adoption-specialist","EdTech district/state adoption + security-compliance process (VPAT, SSO, AI disclosure).",".claude/skills/edtech-adoption-specialist/","\"Help with the district/state EdTech adoption + compliance process\""),

 ("External (plugin)","history-hack-print-qc-auditor","Print-defect / classroom-readiness audit. Available as a plugin — referenced, not vendored in this repo.","(plugin — not in repo)","\"Audit [file] for print defects / classroom readiness\""),
]

wb = openpyxl.Workbook()
ws = wb.active; ws.title = "Master Skills"

thin = Side(style="thin", color="CBD2DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title band
ws.merge_cells("A1:E1")
c = ws["A1"]; c.value = "U.S. History Hack — Master Skills Matrix"
c.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=NAVY); c.alignment = Alignment(vertical="center", indent=1)
ws.row_dimensions[1].height = 30
ws.merge_cells("A2:E2")
c = ws["A2"]; c.value = ("How to use: type  /skill-name  OR just describe the task in the skill's words (the trigger prompt). "
    "Registry of record: .claude/skills/SKILLS.md — one job, one owner. Skills are main-owned; change them only via a skills-only PR.")
c.font = Font(name="Calibri", size=9, italic=True, color="5A6579")
c.fill = PatternFill("solid", fgColor=CREAM); c.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
ws.row_dimensions[2].height = 30

# Header
headers = ["Category", "Skill (say /name)", "The one job it owns", "Where it lives (path)", "Say this to trigger it"]
hr = 3
for j, h in enumerate(headers, 1):
    cell = ws.cell(row=hr, column=j, value=h)
    cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
    cell.border = border
ws.row_dimensions[hr].height = 20

# Rows
r = hr + 1
for cat, skill, job, path, prompt in ROWS:
    newcourse = "NEW-COURSE" in prompt
    vals = [cat, skill, job, path, prompt]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.border = border
        cell.alignment = Alignment(vertical="top", horizontal="left", indent=1, wrap_text=True)
        cell.font = Font(name="Calibri", size=9,
                         color=NAVY if j in (2,) else "2C3446",
                         bold=(j==2))
        if j == 5:
            cell.font = Font(name="Consolas", size=8.5, color=("B22234" if newcourse else "1F3A5F"), bold=newcourse)
        if cat.startswith("★"):
            cell.fill = PatternFill("solid", fgColor=LIGHT)
        elif newcourse:
            cell.fill = PatternFill("solid", fgColor="FBEEDF")
        elif (r % 2) == 0:
            cell.fill = PatternFill("solid", fgColor="FBFCFD")
    ws.row_dimensions[r].height = 44
    r += 1

widths = [20, 34, 62, 40, 52]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w

ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A{hr}:E{r-1}"
ws.sheet_view.showGridLines = False

out = "/tmp/claude-0/-home-user-trooptoteacher-history/241b4294-edc7-58d8-8714-bd87b7a2f4d5/scratchpad/HistoryHack_Master_Skills_Matrix.xlsx"
wb.save(out)
print("saved", out, "| skills:", len(ROWS))
